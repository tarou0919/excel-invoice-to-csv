"""
Excel請求書から情報を抽出するコアモジュール
- 外部YAMLルール対応
- 堅牢なエラー処理
- 詳細なログ出力
"""
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime, date
import re
import logging
from pathlib import Path
from typing import Dict, Any, Optional

logger = logging.getLogger(__name__)


# ============================================================
# テキスト処理ユーティリティ
# ============================================================
def normalize_text(value) -> str:
    """セル値を文字列化して正規化(コロン・空白を除去)"""
    if value is None:
        return ""
    text = str(value).strip()
    # 末尾のコロン・全角コロンを除去
    text = re.sub(r"[::]\s*$", "", text)
    return text


def is_label_match(cell_text, labels) -> bool:
    """セルテキストがラベル候補のいずれかと一致するかチェック"""
    normalized = normalize_text(cell_text)
    if not normalized:
        return False
    for label in labels:
        if normalized == label or normalized.startswith(label):
            return True
    return False


# ============================================================
# 値の取得・変換
# ============================================================
def get_merged_cell_value(ws, cell):
    """結合セルの場合、結合範囲の起点セルの値を取得"""
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            return top_left.value
    return cell.value


def get_merged_range_for_cell(ws, cell):
    """セルが結合セルの一部なら、その結合範囲を返す。違えば None"""
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return merged_range
    return None


def get_neighbor_value(ws, row, col, direction, label_text=None):
    """
    指定セルの近隣から値を取得
    label_text が指定されれば、それと同じ値は採用しない(ラベル自身を返さないため)
    """
    candidates = []
    if direction == "right":
        candidates = [(row, col + i) for i in range(1, 6)]
    elif direction == "below":
        candidates = [(row + i, col) for i in range(1, 4)]
    elif direction == "right_or_below":
        candidates = [(row, col + i) for i in range(1, 6)] + \
                     [(row + i, col) for i in range(1, 4)]

    for r, c in candidates:
        try:
            cell = ws.cell(row=r, column=c)
            value = get_merged_cell_value(ws, cell)
            if value is None:
                continue
            value_str = str(value).strip()
            if not value_str:
                continue
            # ラベル自身と同じ値はスキップ
            if label_text and normalize_text(value) == normalize_text(label_text):
                continue
            return value
        except Exception as e:
            logger.debug(f"Cell access error at ({r},{c}): {e}")
            continue
    return None


def convert_value(value, data_type):
    """データ型に応じて値を変換"""
    if value is None:
        return None

    try:
        if data_type == "number":
            if isinstance(value, (int, float)):
                return float(value)
            text = str(value).replace(",", "").replace("¥", "").replace("円", "").strip()
            return float(text)

        elif data_type == "date":
            if isinstance(value, (datetime, date)):
                return value.strftime("%Y-%m-%d")
            text = str(value).strip()
            # 日本語形式: "2025年10月15日"
            m = re.match(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", text)
            if m:
                y, mo, d = m.groups()
                return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
            # ISO形式: "2025-10-15" or "2025/10/15"
            m = re.match(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
            if m:
                y, mo, d = m.groups()
                return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
            return text

        else:  # text
            text = str(value).strip()
            text = re.sub(r"\s+", " ", text)
            return text
    except (ValueError, TypeError) as e:
        logger.debug(f"Value conversion error ({data_type}): {value} - {e}")
        return None


# ============================================================
# ラベル探索
# ============================================================
def find_value_by_labels(ws, rule):
    """
    シート内をラベル探索して値を取得
    結合セルがラベルになっている場合、結合範囲の外側を探す
    """
    labels = rule["labels"]
    direction = rule["direction"]

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if not is_label_match(cell.value, labels):
                continue

            label_text = str(cell.value)
            # 結合セルの場合、結合範囲の右端/下端から探索開始
            merged = get_merged_range_for_cell(ws, cell)
            if merged:
                start_row = merged.max_row
                start_col = merged.max_col
            else:
                start_row = cell.row
                start_col = cell.column

            value = get_neighbor_value(ws, start_row, start_col, direction, label_text)
            if value is not None:
                return value
    return None


# ============================================================
# フォールバック処理
# ============================================================
def find_customer_by_honorific(ws, keywords):
    """取引先名のフォールバック: 「御中」「様」を含むセルを探す"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value).strip()
            for keyword in keywords:
                if keyword in text:
                    return text
    return None


def find_total_by_max_amount(ws):
    """合計金額のフォールバック: 数値セルの最大値を取得"""
    max_val = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value > 0:
                if max_val is None or cell.value > max_val:
                    max_val = cell.value
    return max_val


# ============================================================
# メイン抽出関数
# ============================================================
def extract_from_file(file_path, rules: Dict[str, Any]) -> Dict[str, Any]:
    """
    1つのExcelファイルから請求書情報を抽出して dict で返す

    Args:
        file_path: Excelファイルのパス
        rules: load_rules() で読み込んだルール辞書

    Returns:
        抽出結果の辞書(失敗時は _error キーにエラー情報)
    """
    file_path = Path(file_path)
    extraction_rules = rules["extraction_rules"]
    csv_columns = rules["csv_columns"]
    fallback = rules["fallback"]

    result = {col: None for col in csv_columns}
    result["ファイル名"] = file_path.name
    result["_error"] = None

    # ファイル存在チェック
    if not file_path.exists():
        msg = f"ファイルが存在しません: {file_path}"
        logger.error(msg)
        result["_error"] = msg
        return result

    # 拡張子チェック
    if file_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        msg = f"対応していない形式です: {file_path.suffix}"
        logger.warning(f"{file_path.name}: {msg}")
        result["_error"] = msg
        return result

    # Excel読込
    try:
        wb = load_workbook(file_path, data_only=True)
    except InvalidFileException as e:
        msg = f"Excelファイルとして読み込めませんでした(壊れている可能性): {e}"
        logger.error(f"{file_path.name}: {msg}")
        result["_error"] = msg
        return result
    except PermissionError:
        msg = "ファイルが他のアプリで開かれています(Excel等を閉じてください)"
        logger.error(f"{file_path.name}: {msg}")
        result["_error"] = msg
        return result
    except Exception as e:
        msg = f"予期しないエラー: {type(e).__name__}: {e}"
        logger.error(f"{file_path.name}: {msg}")
        result["_error"] = msg
        return result

    if not wb.sheetnames:
        msg = "シートが1つもありません"
        logger.warning(f"{file_path.name}: {msg}")
        result["_error"] = msg
        return result

    # 全シートを順番に走査
    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
        except Exception as e:
            logger.warning(f"{file_path.name}: シート '{sheet_name}' のアクセスに失敗: {e}")
            continue

        for field, rule in extraction_rules.items():
            if field == "ファイル名":
                continue
            if result.get(field) is not None:
                continue  # 既に取得済みはスキップ

            try:
                value = find_value_by_labels(ws, rule)
                if value is not None:
                    converted = convert_value(value, rule["data_type"])
                    if converted is not None and converted != "":
                        result[field] = converted
                        logger.debug(f"{file_path.name}: {field} = {converted} (from sheet '{sheet_name}')")
            except Exception as e:
                logger.warning(
                    f"{file_path.name}: '{field}' の抽出中にエラー: {e}"
                )

    # フォールバック処理
    try:
        ws_main = wb[wb.sheetnames[0]]

        # 取引先名: 御中/様検索
        if "取引先名" in csv_columns and not result.get("取引先名"):
            keywords = fallback.get("customer_honorific_keywords", [])
            customer = find_customer_by_honorific(ws_main, keywords)
            if customer:
                result["取引先名"] = customer
                logger.debug(f"{file_path.name}: 取引先名 (fallback) = {customer}")

        # 合計金額: 最大値検索
        if (
            "合計金額" in csv_columns
            and not result.get("合計金額")
            and fallback.get("total_use_max_amount", True)
        ):
            total = find_total_by_max_amount(ws_main)
            if total:
                result["合計金額"] = total
                logger.debug(f"{file_path.name}: 合計金額 (fallback max) = {total}")

        # 合計金額: 小計+消費税
        if (
            "合計金額" in csv_columns
            and not result.get("合計金額")
            and fallback.get("total_calculate_from_subtotal", True)
            and result.get("小計")
            and result.get("消費税")
        ):
            result["合計金額"] = result["小計"] + result["消費税"]
            logger.debug(f"{file_path.name}: 合計金額 (subtotal+tax) = {result['合計金額']}")
    except Exception as e:
        logger.warning(f"{file_path.name}: フォールバック処理中にエラー: {e}")

    return result
